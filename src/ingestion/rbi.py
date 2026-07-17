"""RBI Payment System Indicators — parse both formats, stitch, validate, store.

RBI publishes PSI in two layouts that together span Apr 2004 to present:
  - Old format (sheet "Old Format"): Apr 2004 - Oct 2019
  - New format (sheet "New Format"): Nov 2019 - present

This module finds all PSI files in data/raw/, auto-detects each one's format
by sheet name, resolves columns via header patterns (config-driven, not fixed
positions), parses each into a common schema, and stitches them into one
continuous monthly series.

Pipeline:
  1. Find all files matching the PSI glob.
  2. Compute a combined SHA256; log new-data vs re-run.
  3. For each file: detect format, resolve columns, parse rows.
  4. Concatenate, sort by date, drop any duplicate months (prefer new format).
  5. Run quality checks; save Parquet + CSV; record the hash.
"""

from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger

from src.config import PsiFormatConfig, Settings, load_settings
from src.ingestion.level_splice import BREAK_DATE, splice_series
from src.ingestion.validation import (
    SchemaValidationError,
    check_data_quality,
    combined_hash,
    detect_freshness,
    record_hash,
    resolve_psi_columns,
)


def _safe_float(value) -> float | None:
    """Convert a cell to float, or None for missing / invalid input.

    Handles comma-formatted numbers ('21,703.44') and placeholders ('-', 'NA').
    """
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "-", "NA", "N/A"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


def _detect_format(
    sheet_name: str,
    formats: dict[str, PsiFormatConfig],
) -> tuple[str | None, PsiFormatConfig | None]:
    """Return the (name, config) of the format whose sheet_match is in the sheet name."""
    sn = sheet_name.lower()
    for fmt_name, fmt in formats.items():
        if fmt.sheet_match.lower() in sn:
            return fmt_name, fmt
    return None, None


# ── Individual monthly PSI file parser ────────────────────────────────────
# The RBI scraper downloads one XLSX per month (e.g. PSIAPRIL2026*.XLSX).
# These have a single sheet named "<Month> <Year>" with fixed row layout
# but variable row offsets. Column F holds the current month's value.

_MONTHLY_LABEL_MAP: list[tuple[str, str, str | None]] = [
    # (output_column, label_substring, section_context)
    # section_context disambiguates labels that appear in multiple sections.
    # "part_iii" = rows after "PART III" header (infrastructure), None = first match.
    ("credit_cards_outstanding_lakh",  "1.1 credit cards",   "part_iii"),
    ("debit_cards_outstanding_lakh",   "1.2 debit cards",    "part_iii"),
    ("credit_card_vol_lakh",           "4.1 credit cards",   None),
    ("credit_card_pos_vol_lakh",       "4.1.1 pos",          None),
    ("credit_card_other_vol_lakh",     "4.1.2 others",       None),
    ("debit_card_vol_lakh",            "4.2 debit cards",    None),
    ("debit_card_pos_vol_lakh",        "4.2.1 pos",          None),
    ("debit_card_other_vol_lakh",      "4.2.2 others",       None),
    ("pos_terminals_lakh",             "number of pos terminals", "part_iii"),
    ("bharat_qr_lakh",                 "bharat qr",          "part_iii"),
    ("upi_qr_lakh",                    "upi qr",             "part_iii"),
]

# Value columns need a separate pass because the same row label holds both
# volume (col E or F) and value. In individual files, Part II rows have
# volume in col E and value in col F — but Part III infrastructure rows
# only have volume. We handle this by reading both vol and val from the
# same row using different Excel columns.

_MONTHLY_VALUE_LABEL_MAP: list[tuple[str, str]] = [
    # (output_column, label_substring) — value is read from col J (index 9)
    ("credit_card_val_cr",       "4.1 credit cards"),
    ("credit_card_pos_val_cr",   "4.1.1 pos"),
    ("credit_card_other_val_cr", "4.1.2 others"),
    ("debit_card_val_cr",        "4.2 debit cards"),
    ("debit_card_pos_val_cr",    "4.2.1 pos"),
    ("debit_card_other_val_cr",  "4.2.2 others"),
]

_MONTH_NAMES = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]


def _is_monthly_psi(sheet_name: str) -> bool:
    """True if the sheet name looks like 'April 2026' (individual monthly PSI)."""
    parts = sheet_name.strip().split()
    return (
        len(parts) == 2
        and parts[0].lower() in _MONTH_NAMES
        and parts[1].isdigit()
        and len(parts[1]) == 4
    )


def _parse_monthly_psi(filepath: Path) -> pd.DataFrame:
    """Parse one individual monthly PSI file into a single-row DataFrame.

    Layout: columns C-F = Volume (lakh), columns G-J = Value (₹ crore).
    Column F (idx 5) = current month volume, column J (idx 9) = current month value.
    Row positions vary between files; we match by label text.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    sheet_name = ws.title.strip()

    parts = sheet_name.split()
    month_str, year_str = parts[0], parts[1]
    date = pd.to_datetime(f"{month_str} {year_str}", format="%B %Y")

    vol_col = 5   # column F (0-indexed)
    val_col = 9   # column J (0-indexed)

    all_rows: list[tuple[int, str, list]] = []
    part_iii_start = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=120, max_col=11, values_only=True), start=1):
        label = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if "part iii" in label.lower():
            part_iii_start = r_idx
        all_rows.append((r_idx, label, list(row)))

    record: dict = {"date": date, "source_format": "monthly"}

    for out_col, label_sub, section in _MONTHLY_LABEL_MAP:
        label_sub_lower = label_sub.lower()
        for r_idx, label, row in all_rows:
            if section == "part_iii" and (part_iii_start is None or r_idx < part_iii_start):
                continue
            if label_sub_lower in label.lower():
                record[out_col] = _safe_float(row[vol_col] if vol_col < len(row) else None)
                break
        else:
            record[out_col] = None

    for out_col, label_sub in _MONTHLY_VALUE_LABEL_MAP:
        label_sub_lower = label_sub.lower()
        for r_idx, label, row in all_rows:
            if label_sub_lower in label.lower():
                record[out_col] = _safe_float(row[val_col] if val_col < len(row) else None)
                break
        else:
            record[out_col] = None

    logger.info(
        f"{filepath.name}: parsed monthly format (sheet '{sheet_name}') "
        f"CC={record.get('credit_cards_outstanding_lakh')}, "
        f"DC={record.get('debit_cards_outstanding_lakh')}"
    )
    return pd.DataFrame([record])


def parse_psi_file(filepath: Path, settings: Settings) -> pd.DataFrame:
    """Parse one PSI file, auto-detecting old vs new format by sheet name."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    fmt_name, fmt = _detect_format(ws.title, settings.rbi_psi.formats)
    if fmt is None:
        expected = [f.sheet_match for f in settings.rbi_psi.formats.values()]
        raise SchemaValidationError(
            f"{filepath.name}: sheet '{ws.title}' matches no known format "
            f"(expected sheet name containing one of {expected})."
        )

    logger.info(f"{filepath.name}: detected '{fmt_name}' format (sheet '{ws.title}')")

    rows = list(ws.iter_rows(values_only=True))

    expected_cols = {name: spec.model_dump() for name, spec in fmt.columns.items()}
    column_map = resolve_psi_columns(rows, fmt.label_row, fmt.unit_row, expected_cols)

    date_idx = fmt.date_col
    records: list[dict] = []
    skipped = 0

    for row in rows[fmt.data_start_row:]:
        if len(row) <= date_idx:
            skipped += 1
            continue
        date_val = row[date_idx]
        if date_val is None or not isinstance(date_val, str):
            skipped += 1
            continue
        try:
            parsed_date = pd.to_datetime(date_val, format=settings.rbi_psi.date_format)
        except (ValueError, TypeError):
            skipped += 1
            continue

        record: dict = {"date": parsed_date, "source_format": fmt_name}
        for col_name, col_idx in column_map.items():
            cell = row[col_idx] if col_idx < len(row) else None
            record[col_name] = _safe_float(cell)
        records.append(record)

    df = pd.DataFrame(records)
    if skipped:
        logger.debug(f"  skipped {skipped} non-data rows (titles, footnotes)")
    logger.info(
        f"  parsed {len(df)} rows | "
        f"{df['date'].min():%b %Y} -> {df['date'].max():%b %Y}"
    )
    return df


def run_rbi_ingestion(settings: Settings | None = None) -> pd.DataFrame:
    """Entry point: find PSI files, parse both formats, stitch, validate, save."""
    if settings is None:
        settings = load_settings()

    raw_dir = settings.paths.rbi_psi_dir
    processed_dir = settings.paths.processed_dir

    combined_files = sorted(raw_dir.glob(settings.rbi_psi.file_pattern))
    monthly_files = sorted(raw_dir.glob("PSI*.XLSX"))

    # Exclude monthly files that also match the combined pattern
    combined_names = {f.name for f in combined_files}
    monthly_files = [f for f in monthly_files if f.name not in combined_names]

    # Verify monthly files are actually individual monthly PSI format
    valid_monthly: list[Path] = []
    for fp in monthly_files:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            if _is_monthly_psi(wb.active.title):
                valid_monthly.append(fp)
            wb.close()
        except Exception:
            pass

    files = combined_files + valid_monthly
    if not files:
        raise FileNotFoundError(
            f"No PSI files matching '{settings.rbi_psi.file_pattern}' or 'PSI*.XLSX' in {raw_dir}\n"
            f"Download from RBI DBIE → Statistics → Financial Sector → Payment Systems\n"
            f"and save to data/raw/rbi_psi/."
        )

    logger.info(
        f"Found {len(combined_files)} combined + {len(valid_monthly)} monthly PSI file(s)"
    )

    # Freshness across all input files
    hash_record = processed_dir / ".rbi_psi.sha256"
    current_hash = combined_hash(files)
    if detect_freshness(current_hash, hash_record):
        logger.info(f"NEW data detected (combined hash {current_hash[:12]}...)")
    else:
        logger.info("REPROCESSING existing files (combined hash unchanged)")

    # Parse each file — route monthly files to the dedicated parser
    monthly_set = set(valid_monthly)
    frames: list[pd.DataFrame] = []
    for fp in files:
        if fp in monthly_set:
            frames.append(_parse_monthly_psi(fp))
        else:
            frames.append(parse_psi_file(fp, settings))

    # Prefer combined format over monthly for overlapping months (combined is
    # the official consolidated release). Monthly files are concat'd last, so
    # we sort by date + source_format to ensure combined ("new"/"old") sorts
    # after "monthly", then keep="last" retains the combined version.
    _fmt_priority = {"monthly": 0, "old": 1, "new": 2}
    combined = pd.concat(frames, ignore_index=True)
    combined["_priority"] = combined["source_format"].map(_fmt_priority).fillna(0)
    combined = (
        combined
        .sort_values(["date", "_priority"])
        .drop_duplicates(subset=["date"], keep="last")
        .drop(columns=["_priority"])
        .reset_index(drop=True)
    )

    # Log the stitch boundary if more than one format is present
    if combined["source_format"].nunique() > 1:
        new_start = combined.loc[combined["source_format"] == "new", "date"].min()
        logger.info(f"Stitched old + new formats at {new_start:%b %Y}")

    # Apply Nov-2019 definitional correction if the series spans the break.
    # The RBI changed how it counts cards outstanding in Nov 2019 (non-financial
    # transaction reclassification). This is a measurement change, not a market
    # event. We estimate and remove the level shift so the pre/post series is
    # commensurate. See src/ingestion/level_splice.py for methodology.
    splice_cols = [
        "credit_cards_outstanding_lakh",
        "debit_cards_outstanding_lakh",
    ]
    spans_break = (
        (combined["date"] < BREAK_DATE).any()
        and (combined["date"] >= BREAK_DATE).any()
    )
    if spans_break:
        for col in splice_cols:
            if col not in combined.columns:
                logger.warning(f"Splice: column '{col}' not found, skipping.")
                continue
            if combined[col].isna().all():
                logger.warning(f"Splice: '{col}' is all-null, skipping.")
                continue
            combined, result = splice_series(combined, col)
            logger.info(
                f"Splice '{col}' at {BREAK_DATE.date()}: "
                f"shift={result.additive_shift:+.2f} lakh "
                f"({result.relative_shift_pct:+.2f}%) applied to "
                f"{(combined['date'] < BREAK_DATE).sum()} pre-break rows"
            )
            if result.pre_slope != 0:
                ratio = result.post_slope / result.pre_slope
                if not (0.5 <= ratio <= 2.0):
                    logger.warning(
                        f"Splice '{col}': pre/post slope ratio={ratio:.2f} "
                        f"(pre={result.pre_slope:+.1f}, post={result.post_slope:+.1f}/mo). "
                        f"May not be a pure level shift -- verify visually."
                    )
    else:
        logger.debug("Splice: series does not span Nov-2019 break, skipping.")

    check_data_quality(
        combined,
        max_null_pct=settings.validation.max_null_pct,
        max_date_gap_days=settings.validation.max_date_gap_days,
        min_rows=settings.validation.min_rows,
    )

    csv_path = processed_dir / "rbi_psi_cards.csv"
    parquet_path = processed_dir / "rbi_psi_cards.parquet"
    combined.to_csv(csv_path, index=False)
    combined.to_parquet(parquet_path, index=False)
    logger.info(
        f"Saved {len(combined)} months "
        f"({combined['date'].min():%b %Y} -> {combined['date'].max():%b %Y}) "
        f"to {csv_path.name} and {parquet_path.name}"
    )

    record_hash(hash_record, current_hash)
    return combined


if __name__ == "__main__":
    df = run_rbi_ingestion()
    print("\nFirst 3 months:")
    print(df.head(3).to_string(index=False))
    print("\nLast 3 months:")
    print(df.tail(3).to_string(index=False))
    print("\nRows per source format:")
    print(df["source_format"].value_counts().to_string())
    print("\nNull counts:")
    print(df.isnull().sum().to_string())
