"""
bankwise.py — RBI bank-wise ATM/POS/card statistics ingestion.

Reads RBI's monthly bank-wise sheets and produces a long-format CSV:
    date, bank_raw, bank, bank_category, credit_outstanding, debit_outstanding

Handles three RBI format variants:
    A) pre-Dec 2020       : no Micro ATM / Bharat QR (X1)
    B) Dec 2020 → ~mid-2023: Micro ATM / Bharat QR added (X2, X3, X4, 1..~14)
    C) ~mid-2023 onwards  : restructured "Infrastructure" + "Card Payments..."
                            layout (~sheets 15..41)

Column positions are NEVER hard-coded. We locate the bank-name column and the
two outstanding-count columns by matching against the merged header text.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import xlrd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# xlrd wrapper — makes .xls sheets quack like openpyxl Worksheets
# ---------------------------------------------------------------------------

class _XlrdSheetWrapper:
    """Thin wrapper around xlrd.Sheet so detect_columns / extract_sheet work unchanged.

    Implements only the subset of the openpyxl Worksheet interface that our
    code actually uses: .title, .iter_rows(min_row, max_row, values_only).
    """

    def __init__(self, sheet: xlrd.sheet.Sheet):
        self._s = sheet
        self.title = sheet.name

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        values_only: bool = False,
    ):
        s = self._s
        if max_row is None:
            max_row = s.nrows
        max_row = min(max_row, s.nrows)
        for r in range(min_row - 1, max_row):  # openpyxl is 1-indexed
            row_vals = []
            for c in range(s.ncols):
                cell = s.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row_vals.append(None)
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    # Return as string if it looks like it should be (e.g. bank name col)
                    row_vals.append(cell.value)
                else:
                    row_vals.append(cell.value)
            yield tuple(row_vals)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

HEADER_SEARCH_ROWS = 8        # how many rows from the top to scan for headers
DATA_START_OFFSET = 1         # data starts this many rows after the last header

# Sub-section labels that appear as rows in column C but are NOT banks.
CATEGORY_LABELS = {
    "SCHEDULED COMMERCIAL BANKS",
    "PUBLIC SECTOR BANKS",
    "PRIVATE SECTOR BANKS",
    "FOREIGN BANKS",
    "PAYMENT BANKS",
    "PAYMENTS BANKS",
    "SMALL FINANCE BANKS",
    "REGIONAL RURAL BANKS",
    "COOPERATIVE BANKS",
    "URBAN COOPERATIVE BANKS",
    "STATE COOPERATIVE BANKS",
    "WHITE LABEL ATM OPERATORS",
    "TOTAL",
    "GRAND TOTAL",
}

# Sheets to ignore — summaries / chart data, not monthly snapshots.
SKIP_SHEET_PATTERNS = (
    re.compile(r"^summary", re.I),
    re.compile(r"^chart", re.I),
)

MONTHS = {
    "JAN": 1, "JANUARY": 1, "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3, "APR": 4, "APRIL": 4,
    "MAY": 5, "JUN": 6, "JUNE": 6, "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8, "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10, "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}


# ---------------------------------------------------------------------------
# Bank-name canonicalisation
# ---------------------------------------------------------------------------

def _strip_suffixes(s: str) -> str:
    s = re.sub(r"\b(LTD|LIMITED|LTD\.|PLC|CORPORATION|CORP)\b", "", s)
    s = re.sub(r"\s+", " ", s).strip(" .,-")
    return s


# Manual overrides — names that won't normalise cleanly on their own.
_BANK_ALIASES = {
    "STATE BANK OF INDIA": "State Bank of India",
    "SBI": "State Bank of India",
    "HDFC BANK": "HDFC Bank",
    "ICICI BANK": "ICICI Bank",
    "AXIS BANK": "Axis Bank",
    "KOTAK MAHINDRA BANK": "Kotak Mahindra Bank",
    "INDUSIND BANK": "IndusInd Bank",
    "YES BANK": "Yes Bank",
    "RBL BANK": "RBL Bank",
    "BANK OF BARODA": "Bank of Baroda",
    "BANK OF INDIA": "Bank of India",
    "BANK OF MAHARASHTRA": "Bank of Maharashtra",
    "CANARA BANK": "Canara Bank",
    "CENTRAL BANK OF INDIA": "Central Bank of India",
    "INDIAN BANK": "Indian Bank",
    "INDIAN OVERSEAS BANK": "Indian Overseas Bank",
    "PUNJAB AND SIND BANK": "Punjab and Sind Bank",
    "PUNJAB NATIONAL BANK": "Punjab National Bank",
    "UCO BANK": "UCO Bank",
    "UNION BANK OF INDIA": "Union Bank of India",
    "AMERICAN EXPRESS BANKING CORPORATION": "American Express",
    "AMERICAN EXPRESS BANK": "American Express",
    "AMEX": "American Express",
    "CITIBANK": "Citi Bank",
    "CITI BANK": "Citi Bank",
    "HSBC": "HSBC",
    "STANDARD CHARTERED BANK": "Standard Chartered Bank",
    "DBS INDIA BANK": "DBS India Bank",
    "DEUTSCHE BANK": "Deutsche Bank",
    "BARCLAYS BANK": "Barclays Bank",
    "BANK OF AMERICA": "Bank of America",
    "SBM BANK INDIA": "SBM Bank India",
    "AIRTEL PAYMENTS BANK": "Airtel Payments Bank",
    "FINO PAYMENTS BANK": "Fino Payments Bank",
    "INDIA POST PAYMENTS BANK": "India Post Payments Bank",
    "JIO PAYMENTS BANK": "Jio Payments Bank",
    "NSDL PAYMENTS BANK": "NSDL Payments Bank",
    "PAYTM PAYMENTS BANK": "Paytm Payments Bank",
    "AU SMALL FINANCE BANK": "AU Small Finance Bank",
    "CAPITAL SMALL FINANCE BANK": "Capital Small Finance Bank",
    "FINCARE SMALL FINANCE BANK": "Fincare Small Finance Bank",
    "EQUITAS SMALL FINANCE BANK": "Equitas Small Finance Bank",
    "ESAF SMALL FINANCE BANK": "ESAF Small Finance Bank",
    "JANA SMALL FINANCE BANK": "Jana Small Finance Bank",
    "NORTH EAST SMALL FINANCE BANK": "North East Small Finance Bank",
    "SURYODAY SMALL FINANCE BANK": "Suryoday Small Finance Bank",
    "UJJIVAN SMALL FINANCE BANK": "Ujjivan Small Finance Bank",
    "UNITY SMALL FINANCE BANK": "Unity Small Finance Bank",
    "UTKARSH SMALL FINANCE BANK": "Utkarsh Small Finance Bank",
    "BANDHAN BANK": "Bandhan Bank",
    "CSB BANK": "CSB Bank",
    "CITY UNION BANK": "City Union Bank",
    "DCB BANK": "DCB Bank",
    "DHANALAKSHMI BANK": "Dhanalakshmi Bank",
    "FEDERAL BANK": "Federal Bank",
    "IDBI BANK": "IDBI Bank",
    "IDFC FIRST BANK": "IDFC First Bank",
    "JAMMU AND KASHMIR BANK": "Jammu and Kashmir Bank",
    "J&K BANK": "Jammu and Kashmir Bank",
    "KARNATAKA BANK": "Karnataka Bank",
    "KARUR VYSYA BANK": "Karur Vysya Bank",
    "SOUTH INDIAN BANK": "South Indian Bank",
    "TAMILNAD MERCANTILE BANK": "Tamilnad Mercantile Bank",
}


def canonical_bank(raw: str) -> str:
    """Return a consistent display name for a bank across months."""
    if not isinstance(raw, str):
        return ""
    key = _strip_suffixes(raw.upper())
    if key in _BANK_ALIASES:
        return _BANK_ALIASES[key]
    # Title-case fallback for banks we haven't seen before.
    return raw.strip().title()


# ---------------------------------------------------------------------------
# Date parsing — sheet title varies wildly across months
# ---------------------------------------------------------------------------

_DATE_RE = re.compile(
    r"(?P<m>Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|"
    r"Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|"
    r"Nov(?:ember)?|Dec(?:ember)?)[\s\-,]+(?P<y>\d{4}|\d{2})(?!\d)",
    re.I,
)


def parse_sheet_date(ws: Worksheet) -> pd.Timestamp:
    """Pull the month/year out of the top rows of the sheet."""
    for row in ws.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS, values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            m = _DATE_RE.search(cell)
            if m:
                month = MONTHS[m.group("m").upper()]
                y = int(m.group("y"))
                # Two-digit years: RBI started in 2011, so anything <50 is 20xx.
                if y < 100:
                    y += 2000 if y < 50 else 1900
                return pd.Timestamp(year=y, month=month, day=1)
    raise ValueError(f"Could not find a date string in sheet '{ws.title}'")


# ---------------------------------------------------------------------------
# Header detection — the core of the fix
# ---------------------------------------------------------------------------

@dataclass
class ColumnMap:
    bank_col: int             # 1-indexed
    credit_outstanding_col: int
    debit_outstanding_col: int
    header_end_row: int       # last row that is header text, not data
    # Infrastructure (optional — present in all formats from 2011)
    atm_onsite_col:   int | None = None
    atm_offsite_col:  int | None = None
    pos_col:          int | None = None   # Format C: single col; Format A: online PoS
    pos_offline_col:  int | None = None   # Format A only: offline PoS (sum with pos_col)
    micro_atm_col:    int | None = None
    bharat_qr_col:    int | None = None
    upi_qr_col:       int | None = None
    # CC transaction volumes (optional)
    cc_pos_vol_col:      int | None = None
    cc_online_vol_col:   int | None = None
    cc_others_vol_col:   int | None = None
    cc_atm_cash_vol_col: int | None = None
    # DC transaction volumes (optional)
    dc_pos_vol_col:      int | None = None
    dc_online_vol_col:   int | None = None
    dc_others_vol_col:   int | None = None
    dc_atm_cash_vol_col: int | None = None


def _stringify(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", " ").strip().lower()


def _scan_grid(ws: Worksheet) -> tuple[list[list[str]], int]:
    """Read the top of the sheet into a 2-D string grid (forgives None)."""
    grid: list[list[str]] = []
    max_col = 0
    for row in ws.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS, values_only=True):
        row_l = [_stringify(v) for v in row]
        grid.append(row_l)
        # find rightmost non-empty
        for i in range(len(row_l) - 1, -1, -1):
            if row_l[i]:
                max_col = max(max_col, i + 1)
                break
    # pad rows to max_col
    for r in grid:
        while len(r) < max_col:
            r.append("")
    return grid, max_col


def _forward_fill(grid: list[list[str]]) -> list[list[str]]:
    """RBI merges header cells — forward-fill blanks within each header row."""
    filled = [r[:] for r in grid]
    for r in filled:
        last = ""
        for j in range(len(r)):
            if r[j]:
                last = r[j]
            else:
                r[j] = last
    return filled


def detect_columns(ws: Worksheet) -> ColumnMap:
    """
    Locate bank-name, credit-outstanding, and debit-outstanding columns by
    matching the joined header text in each column against known signatures.
    Raises ValueError loudly if any column can't be found.
    """
    grid, max_col = _scan_grid(ws)
    if max_col == 0:
        raise ValueError(f"Sheet '{ws.title}' looks empty in the top {HEADER_SEARCH_ROWS} rows")

    filled = _forward_fill(grid)

    # Find the bank-name column: any header row says exactly 'bank name'
    bank_col = None
    header_end_row = 0
    for i, row in enumerate(filled):
        for j, v in enumerate(row):
            if v.strip() == "bank name":
                bank_col = j + 1
                header_end_row = max(header_end_row, i + 1)
                break
        if bank_col:
            break
    if bank_col is None:
        raise ValueError(f"Sheet '{ws.title}': no 'Bank Name' column found")

    # For each column, build a joined header signature from rows below the
    # group label down through the last header row.
    def col_signature(j: int) -> str:
        return " | ".join(r[j] for r in filled if r[j])

    # Identify the "Credit Cards" and "Debit Cards" group spans, then within
    # each span find the column whose lower-level header says "outstanding".
    def find_outstanding(group_label_substr: str) -> tuple[int, int]:
        """Return (col_1_indexed, last_header_row_1_indexed)."""
        # First locate the group header row (the row containing the group label)
        group_row = None
        group_cols: list[int] = []
        for i, row in enumerate(filled):
            for j, v in enumerate(row):
                if group_label_substr in v:
                    if group_row is None:
                        group_row = i
                    if i == group_row:
                        group_cols.append(j)
        if group_row is None:
            raise ValueError(
                f"Sheet '{ws.title}': could not find '{group_label_substr}' group header"
            )

        # Forward-fill spreads the label across the merged span — find the run.
        start = min(group_cols)
        end = start
        for j in range(start, max_col):
            if filled[group_row][j] == filled[group_row][start]:
                end = j
            else:
                break

        # In rows below group_row, look for "outstanding" within [start..end].
        last_header_row = group_row + 1
        outstanding_col = None
        for i in range(group_row + 1, len(filled)):
            for j in range(start, end + 1):
                cell = filled[i][j]
                if "outstanding" in cell and "card" not in cell.replace(
                    "outstanding cards", ""
                ).strip(" ,"):
                    # Generic: cell mentions 'outstanding'
                    pass
                if "outstanding" in cell:
                    outstanding_col = j + 1
                    last_header_row = max(last_header_row, i + 1)
                    break
            if outstanding_col:
                break

        if outstanding_col is None:
            # Format C: the "outstanding" word lives one row above the group
            # label, in an upper "Number - Outstanding" banner. Fall back to
            # column where the raw label is "credit cards" / "debit cards"
            # directly under the banner.
            wanted = group_label_substr.replace(" cards", "").strip()  # "credit"/"debit"
            wanted_full = f"{wanted} cards"
            for i in range(len(filled)):
                for j in range(max_col):
                    if filled[i][j].strip() == wanted_full:
                        # Confirm this column lives under a "Number - Outstanding"
                        # banner somewhere above.
                        banner_ok = any(
                            "outstanding" in filled[k][j]
                            for k in range(i)
                        )
                        if banner_ok:
                            outstanding_col = j + 1
                            last_header_row = max(last_header_row, i + 1)
                            break
                if outstanding_col:
                    break

        if outstanding_col is None:
            raise ValueError(
                f"Sheet '{ws.title}': could not find 'outstanding' column inside "
                f"'{group_label_substr}' group (cols {start + 1}..{end + 1})"
            )
        return outstanding_col, last_header_row

    credit_col, cr_hdr_end = find_outstanding("credit cards")
    debit_col, db_hdr_end = find_outstanding("debit cards")
    header_end_row = max(header_end_row, cr_hdr_end, db_hdr_end)

    # Sanity: also catch the Format-C banner row "Number - Outstanding".
    # If we see it below header_end_row, push header_end_row down past it.
    for i, row in enumerate(filled):
        if any("number - outstanding" in v or "no. of outstanding" in v for v in row):
            header_end_row = max(header_end_row, i + 1)

    # ── Extra columns (Format A 2011-2021 and Format B/C 2020+) ─────────────
    # Build a per-column signature by joining all non-empty header cell texts.
    # All extra cols are optional: if not found we return None and the sheet
    # is still processed normally.
    #
    # Format A (2011-2021): ATMs | PoS(online) | PoS(offline) | CC outstanding |
    #   CC txn ATM | CC txn PoS | DC outstanding | DC txn ATM | DC txn PoS
    # Format C (2022+): ATMs | PoS | Micro ATM | Bharat QR | UPI QR |
    #   CC outstanding | DC outstanding | CC PoS/Online/Others/Cash | DC same

    def col_sig(j: int) -> str:
        return " | ".join(r[j] for r in filled if j < len(r) and r[j])

    def find_col(*keyword_sets) -> int | None:
        """Return 1-indexed column whose signature matches ANY of the keyword sets.
        Each set is a tuple of strings that must ALL be present (AND within set,
        OR between sets).
        """
        sets = [keyword_sets] if isinstance(keyword_sets[0], str) else list(keyword_sets)
        for j in range(max_col):
            sig = col_sig(j)
            for kws in sets:
                if all(kw in sig for kw in kws):
                    return j + 1
        return None

    # ATMs — same keyword in both formats
    atm_onsite  = find_col(("atms", "on-site"),  ("atms & crms", "on-site"))
    atm_offsite = find_col(("atms", "off-site"), ("atms & crms", "off-site"))

    # PoS terminals
    # Format A: two cols (online + offline) — we take online as primary,
    #   offline stored separately and summed in extract_sheet
    # Format C: single col under infrastructure
    pos_col         = find_col(("pos", "on-line"), ("infrastructure", "pos", "number - outstanding"))
    pos_offline_col = find_col(("pos", "off-line"),)   # Format A only; None in Format C

    # Format B/C only
    micro_atm = find_col(("micro atms",),)
    bharat_qr = find_col(("bharat qr",),)
    upi_qr    = find_col(("upi qr",),)

    # CC transaction volume columns
    # Format A: "credit cards" + "pos" + "transactions (actuals)"
    # Format C: "credit card" + "card payments" + "at pos" + "volume"
    cc_pos_vol = find_col(
        ("credit card", "card payments", "at pos", "volume"),
        ("credit cards", "pos", "transactions (actuals)"),
        ("credit cards", "pos", "no. of transactions"),
    )
    cc_online_vol   = find_col(("credit card", "online", "volume"),)
    cc_others_vol   = find_col(("credit card", "card payments", "others", "volume"),)
    # Format A: CC ATM txn = cash-like (ATM usage of CC)
    cc_atm_cash_vol = find_col(
        ("credit card", "cash withdrawal", "volume"),
        ("credit cards", "atm", "transactions (actuals)"),
        ("credit cards", "atm", "no. of transactions"),
    )

    # DC transaction volume columns
    dc_pos_vol = find_col(
        ("debit card", "card payments", "at pos", "volume"),
        ("debit cards", "pos", "transactions (actuals)"),
        ("debit cards", "pos", "no. of transactions"),
    )
    dc_online_vol   = find_col(("debit card", "online", "volume"),)
    dc_others_vol   = find_col(("debit card", "card payments", "others", "volume"),)
    dc_atm_cash_vol = find_col(
        ("debit card", "cash withdrawal", "atm", "volume"),
        ("debit cards", "atm", "transactions (actuals)"),
        ("debit cards", "atm", "no. of transactions"),
    )

    return ColumnMap(
        bank_col=bank_col,
        credit_outstanding_col=credit_col,
        debit_outstanding_col=debit_col,
        header_end_row=header_end_row,
        atm_onsite_col=atm_onsite,
        atm_offsite_col=atm_offsite,
        pos_col=pos_col,
        pos_offline_col=pos_offline_col,
        micro_atm_col=micro_atm,
        bharat_qr_col=bharat_qr,
        upi_qr_col=upi_qr,
        cc_pos_vol_col=cc_pos_vol,
        cc_online_vol_col=cc_online_vol,
        cc_others_vol_col=cc_others_vol,
        cc_atm_cash_vol_col=cc_atm_cash_vol,
        dc_pos_vol_col=dc_pos_vol,
        dc_online_vol_col=dc_online_vol,
        dc_others_vol_col=dc_others_vol,
        dc_atm_cash_vol_col=dc_atm_cash_vol,
    )


# ---------------------------------------------------------------------------
# Per-sheet extraction
# ---------------------------------------------------------------------------

def _is_number(v) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        try:
            float(s)
            return True
        except ValueError:
            return False
    return False


def _to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def extract_sheet(ws: Worksheet, verbose: bool = False) -> pd.DataFrame:
    date = parse_sheet_date(ws)
    cols = detect_columns(ws)

    if verbose:
        print(
            f"  {ws.title:>4} | {date.strftime('%Y-%m')} | "
            f"bank=col{cols.bank_col} cr=col{cols.credit_outstanding_col} "
            f"db=col{cols.debit_outstanding_col} hdr_end=row{cols.header_end_row}"
        )

    records: list[dict] = []
    current_category: str | None = None
    start_row = cols.header_end_row + DATA_START_OFFSET

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        raw_bank = row[cols.bank_col - 1] if len(row) >= cols.bank_col else None
        if not isinstance(raw_bank, str):
            continue
        name = raw_bank.strip()
        if not name:
            continue

        upper = name.upper().strip(" :")
        # Category header row?
        if upper in CATEGORY_LABELS or any(
            upper == c or upper.startswith(c + " ") for c in CATEGORY_LABELS
        ):
            if upper not in {"TOTAL", "GRAND TOTAL"}:
                current_category = upper.title()
            continue

        # Footnote-ish rows: row's data columns are all empty/non-numeric
        cr_val = _to_number(row[cols.credit_outstanding_col - 1]) \
            if len(row) >= cols.credit_outstanding_col else None
        db_val = _to_number(row[cols.debit_outstanding_col - 1]) \
            if len(row) >= cols.debit_outstanding_col else None
        if cr_val is None and db_val is None:
            continue

        def _get(col_idx):
            if col_idx is None:
                return None
            return _to_number(row[col_idx - 1]) if len(row) >= col_idx else None

        # PoS: Format A has online+offline; sum them. Format C has single col.
        pos_online  = _get(cols.pos_col)
        pos_offline = _get(cols.pos_offline_col)
        if pos_online is not None and pos_offline is not None:
            pos_total = pos_online + pos_offline
        else:
            pos_total = pos_online  # Format C or None

        records.append({
            "date":               date,
            "sheet":              ws.title,
            "bank_raw":           name,
            "bank":               canonical_bank(name),
            "bank_category":      current_category,
            "credit_outstanding": cr_val,
            "debit_outstanding":  db_val,
            # Infrastructure
            "atm_onsite":         _get(cols.atm_onsite_col),
            "atm_offsite":        _get(cols.atm_offsite_col),
            "pos_terminals":      pos_total,
            "micro_atm":          _get(cols.micro_atm_col),
            "bharat_qr":          _get(cols.bharat_qr_col),
            "upi_qr":             _get(cols.upi_qr_col),
            # CC transaction volumes
            "cc_pos_vol":         _get(cols.cc_pos_vol_col),
            "cc_online_vol":      _get(cols.cc_online_vol_col),
            "cc_others_vol":      _get(cols.cc_others_vol_col),
            "cc_atm_cash_vol":    _get(cols.cc_atm_cash_vol_col),
            # DC transaction volumes
            "dc_pos_vol":         _get(cols.dc_pos_vol_col),
            "dc_online_vol":      _get(cols.dc_online_vol_col),
            "dc_others_vol":      _get(cols.dc_others_vol_col),
            "dc_atm_cash_vol":    _get(cols.dc_atm_cash_vol_col),
        })

    return pd.DataFrame.from_records(records)


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def should_skip_sheet(name: str) -> bool:
    return any(p.search(name) for p in SKIP_SHEET_PATTERNS)


def _open_sheets(filepath: Path, verbose: bool = False):
    """Open a workbook and yield (sheet_name, worksheet) pairs.

    Handles both .xlsx (openpyxl) and .xls (xlrd) formats transparently.
    For .xls files, yields _XlrdSheetWrapper instances that implement the
    same iter_rows / .title interface as openpyxl Worksheets.
    """
    ext = filepath.suffix.lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(str(filepath))
        if verbose:
            print(f"Loaded {filepath.name} (xls) -- {wb.nsheets} sheets")
        for idx in range(wb.nsheets):
            ws = _XlrdSheetWrapper(wb.sheet_by_index(idx))
            yield ws.title, ws
    else:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        if verbose:
            print(f"Loaded {filepath.name} -- {len(wb.sheetnames)} sheets")
        for name in wb.sheetnames:
            yield name, wb[name]


def ingest(filepath: Path, verbose: bool = False) -> pd.DataFrame:
    """Ingest one bankwise file (.xlsx or .xls). Returns long-format DataFrame."""
    if not filepath.exists():
        raise FileNotFoundError(filepath)

    frames: list[pd.DataFrame] = []
    errors: list[tuple[str, str]] = []

    for name, ws in _open_sheets(filepath, verbose=verbose):
        if should_skip_sheet(name):
            if verbose:
                print(f"  skip: {name}")
            continue
        try:
            df = extract_sheet(ws, verbose=verbose)
            if df.empty:
                errors.append((name, "no data rows extracted"))
                continue
            frames.append(df)
        except Exception as e:  # noqa: BLE001 — we want loud, per-sheet failure
            errors.append((name, str(e)))
            if verbose:
                print(f"  ERROR on '{name}': {e}", file=sys.stderr)

    if errors:
        print(f"\n{len(errors)} sheet(s) failed:", file=sys.stderr)
        for n, msg in errors:
            print(f"  {n}: {msg}", file=sys.stderr)

    if not frames:
        raise RuntimeError(f"No sheets ingested from {filepath.name}")

    out = pd.concat(frames, ignore_index=True)
    out = out.sort_values(["date", "bank"]).reset_index(drop=True)
    return out


# ---------------------------------------------------------------------------
# Known data quality issues — imputation rules
# ---------------------------------------------------------------------------

# Entries with implausibly small values that should be imputed from trend.
# Key: (bank_canonical, date_str), Value: description for logging.
_KNOWN_DATA_ERRORS: dict[tuple[str, str], str] = {
    ("HDFC Bank", "2025-05-01"): (
        "RBI Sheet 41 May-2025: credit_outstanding = 6 (should be ~24M). "
        "Data entry error. Imputed from Apr-2025 value + trailing 3-month trend."
    ),
}


def _impute_known_errors(df: pd.DataFrame) -> pd.DataFrame:
    """Replace known RBI data entry errors with trend-imputed values.

    Only operates on entries registered in _KNOWN_DATA_ERRORS.
    Logs every imputation. Never imputes silently.
    """
    df = df.copy()
    for (bank_key, date_str), description in _KNOWN_DATA_ERRORS.items():
        date = pd.Timestamp(date_str)
        mask = (df["bank"] == bank_key) & (df["date"] == date)
        if not mask.any():
            continue

        # Compute imputed value from trailing 3-month trend
        bank_hist = (
            df[(df["bank"] == bank_key) & (df["date"] < date)]
            .sort_values("date")
            .tail(3)
        )
        if len(bank_hist) < 2:
            print(f"  IMPUTE WARNING: {bank_key} {date.date()} — not enough history", file=sys.stderr)
            continue

        cr_vals = bank_hist["credit_outstanding"].dropna()
        if len(cr_vals) >= 2:
            avg_delta = cr_vals.diff().dropna().mean()
            imputed_cr = float(cr_vals.iloc[-1] + avg_delta)
            original = df.loc[mask, "credit_outstanding"].iloc[0]
            df.loc[mask, "credit_outstanding"] = imputed_cr
            print(
                f"  IMPUTED: {bank_key} {date.date()} credit_outstanding: "
                f"{original} -> {imputed_cr:,.0f} ({description})"
            )

        db_vals = bank_hist["debit_outstanding"].dropna()
        if len(db_vals) >= 2:
            avg_delta_db = db_vals.diff().dropna().mean()
            imputed_db = float(db_vals.iloc[-1] + avg_delta_db)
            original_db = df.loc[mask, "debit_outstanding"].iloc[0]
            # Only impute debit if it also looks wrong (< 1% of previous)
            if original_db is not None and original_db < db_vals.iloc[-1] * 0.01:
                df.loc[mask, "debit_outstanding"] = imputed_db
                print(
                    f"  IMPUTED: {bank_key} {date.date()} debit_outstanding: "
                    f"{original_db} -> {imputed_db:,.0f}"
                )

    return df


# ---------------------------------------------------------------------------
# Full ingestion pipeline
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_RAW_DIR      = _PROJECT_ROOT / "data" / "raw" / "rbi_bankwise"
_PROCESSED    = _PROJECT_ROOT / "data" / "processed"


def run_bankwise_ingestion(settings=None, verbose: bool = False) -> pd.DataFrame:
    """Entry point: find all bankwise source files, parse, deduplicate, save.

    Sources (in priority order for deduplication):
      1. RBI_Data_Debit_Credit_1.xlsx — consolidated file with monthly sheets
         (Format C, Jan 2022 - May 2025) and summary sheets.
      2. Year-folder files (2011/ ... 2021/) — individual monthly .xls/.xlsx

    Outputs:
      data/processed/bankwise_cards_cc.parquet — credit card outstanding
      data/processed/bankwise_cards_dc.parquet — debit card outstanding
    """
    _PROCESSED.mkdir(parents=True, exist_ok=True)

    frames: list[pd.DataFrame] = []
    source_tags: list[str] = []

    # Source 1: Consolidated XLSX
    consolidated = _RAW_DIR / "RBI_Data_Debit_Credit_1.xlsx"
    if consolidated.exists():
        print(f"Processing consolidated file: {consolidated.name}")
        df_consolidated = ingest(consolidated, verbose=verbose)
        df_consolidated["source"] = "xlsx_consolidated"
        frames.append(df_consolidated)
        source_tags.append(f"consolidated: {len(df_consolidated)} rows")

    # Source 2: Year-folder files
    year_dirs = sorted(d for d in _RAW_DIR.iterdir() if d.is_dir() and d.name.isdigit())
    for year_dir in year_dirs:
        # Deduplicate: Windows glob is case-insensitive, so *.xlsx and *.XLSX
        # return the same files. Use a set of lowercased stems to avoid doubles.
        seen: set[str] = set()
        all_files: list[Path] = []
        for f in sorted(year_dir.iterdir()):
            if f.suffix.lower() in (".xls", ".xlsx") and f.name.lower() not in seen:
                seen.add(f.name.lower())
                all_files.append(f)
        print(f"  {year_dir.name}/: {len(all_files)} files")
        for f in all_files:
            try:
                df_file = ingest(f, verbose=verbose)
                ext = f.suffix.lower()
                df_file["source"] = f"{'xlsx' if ext == '.xlsx' else 'xls'}_monthly"
                frames.append(df_file)
            except Exception as e:
                print(f"  WARNING: {f.name}: {e}", file=sys.stderr)

    if not frames:
        raise FileNotFoundError(
            f"No bankwise source files found in {_RAW_DIR}. "
            f"Expected RBI_Data_Debit_Credit_1.xlsx and/or year folders."
        )

    combined = pd.concat(frames, ignore_index=True)

    # Deduplicate: prefer xlsx_consolidated over year-folder data
    # Sort by source priority (consolidated first), then drop duplicates per (date, bank)
    source_priority = {"xlsx_consolidated": 0, "xlsx_monthly": 1, "xls_monthly": 2}
    combined["_priority"] = combined["source"].map(source_priority).fillna(9)
    combined = (
        combined.sort_values(["date", "bank", "_priority"])
        .drop_duplicates(subset=["date", "bank"], keep="first")
        .drop(columns=["_priority"])
        .sort_values(["date", "bank"])
        .reset_index(drop=True)
    )

    # Impute known data entry errors
    combined = _impute_known_errors(combined)

    # Rename to bank_name for downstream compatibility
    combined = combined.rename(columns={"bank": "bank_name", "bank_raw": "bank_name_raw"})

    # Flag low-coverage banks (< 12 months of credit data)
    coverage = combined.groupby("bank_name")["credit_outstanding"].apply(
        lambda s: s.notna().sum()
    )
    low_banks = set(coverage[coverage < 12].index)
    combined["low_coverage"] = combined["bank_name"].isin(low_banks)

    # Split into CC and DC parquets — include extra columns where available
    _infra = ["atm_onsite", "atm_offsite", "pos_terminals",
              "micro_atm", "bharat_qr", "upi_qr"]
    _base  = ["date", "bank_name", "bank_name_raw", "bank_category", "source", "low_coverage"]

    _cc_extra = ["cc_pos_vol", "cc_online_vol", "cc_others_vol", "cc_atm_cash_vol"]
    _dc_extra = ["dc_pos_vol", "dc_online_vol", "dc_others_vol", "dc_atm_cash_vol"]

    # Only keep extra cols that actually exist (older-format files produce all-NaN cols)
    def _keep(cols_list):
        return [c for c in cols_list if c in combined.columns and combined[c].notna().any()]

    cc_cols = _base + ["credit_outstanding"] + _keep(_infra) + _keep(_cc_extra)
    dc_cols = _base + ["debit_outstanding"]  + _keep(_infra) + _keep(_dc_extra)

    cc = combined[cc_cols].rename(columns={"credit_outstanding": "cc_outstanding"})
    dc = combined[dc_cols].rename(columns={"debit_outstanding": "dc_outstanding"})

    cc_path = _PROCESSED / "bankwise_cards_cc.parquet"
    dc_path = _PROCESSED / "bankwise_cards_dc.parquet"
    cc_csv  = _PROCESSED / "bankwise_cards_cc.csv"
    dc_csv  = _PROCESSED / "bankwise_cards_dc.csv"

    cc.to_parquet(cc_path, index=False)
    cc.to_csv(cc_csv, index=False)
    dc.to_parquet(dc_path, index=False)
    dc.to_csv(dc_csv, index=False)

    n_banks = combined["bank_name"].nunique()
    date_range = f"{combined['date'].min().date()} -> {combined['date'].max().date()}"
    print(
        f"\nBankwise ingestion complete:\n"
        f"  {len(combined):,} rows | {n_banks} banks | {date_range}\n"
        f"  CC: {cc_path.name} ({len(cc):,} rows)\n"
        f"  DC: {dc_path.name} ({len(dc):,} rows)\n"
        f"  Sources: {', '.join(source_tags)}"
    )

    return combined


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", type=Path, default=None,
                   help="Path to single RBI bank-wise xlsx (legacy mode)")
    p.add_argument("--out", type=Path, default=None)
    p.add_argument("-v", "--verbose", action="store_true")
    args = p.parse_args()

    if args.xlsx:
        # Legacy single-file mode
        df = ingest(args.xlsx, verbose=args.verbose)
        out = args.out or Path("bankwise_cards.csv")
        df.to_csv(out, index=False)
        print(
            f"\nWrote {len(df):,} rows -> {out}\n"
            f"  date range : {df['date'].min().date()} -> {df['date'].max().date()}\n"
            f"  banks      : {df['bank'].nunique()} unique\n"
            f"  sheets in  : {df['sheet'].nunique()}\n"
        )
    else:
        # Full pipeline mode
        run_bankwise_ingestion(verbose=args.verbose)
    return 0


if __name__ == "__main__":
    sys.exit(main())