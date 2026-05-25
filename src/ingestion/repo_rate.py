"""RBI Repo Rate — parse event-based table, forward-fill to monthly series, save.

Source file: RepoRate2007.XLSX (RBI Handbook of Statistics Table 43).
Two sheets:
  T_43(I)  — 2007-01-31 to 2015-09-29  (35 change events)
  T_43(II) — 2016-04-05 to 2025-02-07  (20 change events)

The repo rate only changes when the RBI Monetary Policy Committee meets
(roughly every 2 months). The source records each change date and the new
rate. This pipeline:
  1. Parses both sheets into a list of (effective_date, repo_rate) events.
  2. Handles the gap between Oct 2015 and Apr 2016 (7 months) by bridging
     from the last known rate in T_43(I) to the first in T_43(II).
  3. Forward-fills to a continuous monthly series (month-start dates).
  4. Validates coverage and saves.

Gap note (Oct 2015 – Mar 2016):
  The RBI held the repo rate at 6.75% from Sep 2015 through Apr 2016, when
  it was cut to 6.5%. The source file has a structural break across the two
  sheets but the rate was unchanged during this period. The gap is bridged
  automatically by forward-fill.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger

from src.config import Settings, load_settings
from src.ingestion.validation import SchemaValidationError, check_data_quality

# Column index of the Repo Rate in Table 43 (0-indexed within the row tuple)
# Layout: col0=blank, col1=Effective Date, col2=Bank Rate, col3=Repo, col4=Reverse Repo
_REPO_COL = 3


def _parse_table43_sheet(ws) -> list[tuple[datetime.datetime, float]]:
    """Extract (effective_date, repo_rate) tuples from one Table 43 sheet.

    Skips rows where the repo column is blank or '-' (no change that day).
    """
    rows = list(ws.iter_rows(values_only=True))
    events: list[tuple[datetime.datetime, float]] = []

    for row in rows:
        if len(row) <= _REPO_COL:
            continue

        date_cell = row[1]
        repo_cell = row[_REPO_COL]

        # Date cell must be a datetime
        if not isinstance(date_cell, datetime.datetime):
            continue

        # Repo cell must be a numeric-looking value
        if repo_cell is None or str(repo_cell).strip() in ("", "-", "3"):
            continue

        try:
            rate = float(str(repo_cell).strip())
        except (ValueError, TypeError):
            continue

        events.append((date_cell, rate))

    return events


def _build_monthly_series(
    events: list[tuple[datetime.datetime, float]],
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DataFrame:
    """Forward-fill repo rate events into a monthly (month-start) series.

    Args:
        events:     sorted list of (effective_date, rate) change events
        start_date: first month to include (month-start)
        end_date:   last month to include (month-start)

    Returns:
        DataFrame with columns: date, repo_rate
    """
    if not events:
        raise SchemaValidationError("No repo rate events parsed — check source file.")

    # Build monthly date range
    months = pd.date_range(start=start_date, end=end_date, freq="MS")

    # Sort events chronologically
    events_sorted = sorted(events, key=lambda x: x[0])

    monthly_rates: list[float | None] = []
    last_rate: float | None = None

    for month in months:
        # Apply all events whose effective date falls on or before this month
        for ev_date, ev_rate in events_sorted:
            if ev_date <= month + pd.offsets.MonthEnd(0):  # within the month
                last_rate = ev_rate
        monthly_rates.append(last_rate)

    df = pd.DataFrame({"date": months, "repo_rate": monthly_rates})

    # Warn if there are leading nulls (months before the first known rate)
    null_leading = df["repo_rate"].isna().sum()
    if null_leading > 0:
        logger.warning(
            f"{null_leading} months at the start of the series have no repo rate "
            "(before the first known change event). They will be null in the output."
        )

    return df


def run_repo_rate_ingestion(settings: Settings | None = None) -> pd.DataFrame:
    """Entry point: parse both Table 43 sheets, merge, forward-fill, save."""
    if settings is None:
        settings = load_settings()

    raw_dir = settings.paths.rbi_repo_dir
    processed_dir = settings.paths.processed_dir

    candidates = sorted(
        list(raw_dir.glob(settings.rbi_repo.file_pattern)) +
        list(raw_dir.glob(settings.rbi_repo.file_pattern.replace(".xlsx", ".XLSX")))
    )
    if not candidates:
        raise FileNotFoundError(
            f"No repo rate file matching '{settings.rbi_repo.file_pattern}' in {raw_dir}.\n"
            "Copy RepoRate2007.XLSX into data/raw/rbi_repo_rate/ and retry.\n"
            "Source: RBI Handbook of Statistics → Table 43 (Policy Rates)"
        )
    filepath = candidates[-1]
    logger.info(f"Parsing repo rate file: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    logger.info(f"Sheets found: {wb.sheetnames}")

    all_events: list[tuple[datetime.datetime, float]] = []
    for sheet_name in wb.sheetnames:
        if "43" in sheet_name:
            ws = wb[sheet_name]
            sheet_events = _parse_table43_sheet(ws)
            logger.info(f"  Sheet '{sheet_name}': {len(sheet_events)} rate change events")
            if sheet_events:
                logger.debug(f"    First: {sheet_events[0]}, Last: {sheet_events[-1]}")
            all_events.extend(sheet_events)
        else:
            logger.warning(f"  Sheet '{sheet_name}' not recognised — skipping.")

    if not all_events:
        raise SchemaValidationError(
            "No repo rate events found in any sheet. "
            "Check that the source file is the RBI Handbook Table 43 format."
        )

    all_events.sort(key=lambda x: x[0])
    logger.info(
        f"Total: {len(all_events)} rate change events | "
        f"{all_events[0][0].strftime('%b %Y')} → {all_events[-1][0].strftime('%b %Y')}"
    )

    # Build monthly series starting from April 2004 (aligns with RBI PSI series)
    # End at the month of the last event (or current month, whichever is later)
    series_start = pd.Timestamp("2004-04-01")
    series_end = pd.Timestamp(all_events[-1][0]).replace(day=1) + pd.offsets.MonthEnd(3)
    series_end = series_end.replace(day=1)  # next quarter end → normalise to month-start

    logger.info(
        f"Building monthly series: {series_start:%b %Y} → {series_end:%b %Y}"
    )

    df = _build_monthly_series(all_events, series_start, series_end)

    # Warn about the known Oct 2015 – Mar 2016 structural gap
    gap_months = df[
        (df["date"] >= "2015-10-01") & (df["date"] <= "2016-03-01")
    ]
    if gap_months["repo_rate"].isna().any():
        logger.warning(
            "Gap detected in Oct 2015 – Mar 2016 (between T_43(I) and T_43(II)). "
            "RBI held the rate at 6.75% during this period — verify the forward-fill "
            "is correct and update settings.toml structural_events if needed."
        )
    else:
        logger.info(
            "Oct 2015 – Mar 2016 gap bridged cleanly by forward-fill "
            f"(rate = {gap_months['repo_rate'].iloc[0]}%)."
        )

    check_data_quality(
        df,
        date_col="date",
        max_null_pct=15.0,  # Apr 2004–Dec 2006 pre-dates first known rate
        max_date_gap_days=35,
        min_rows=100,
    )

    csv_path = processed_dir / "repo_rate.csv"
    parquet_path = processed_dir / "repo_rate.parquet"
    df.to_csv(csv_path, index=False)
    df.to_parquet(parquet_path, index=False)
    logger.info(f"Saved {csv_path.name} and {parquet_path.name}")

    # Print a few key rate changes for sanity check
    logger.info("Key rate change months (for sanity check):")
    changes = df[df["repo_rate"].diff().abs() > 0.01]
    for _, row in changes.iterrows():
        logger.info(f"  {row['date']:%b %Y}: {row['repo_rate']:.2f}%")

    return df


if __name__ == "__main__":
    df = run_repo_rate_ingestion()
    print(f"\nShape: {df.shape}")
    print(df.tail(8).to_string(index=False))
    print(f"\nRate change events (non-zero month-on-month diff):")
    changes = df[df["repo_rate"].diff().abs() > 0.01]
    print(changes.to_string(index=False))
