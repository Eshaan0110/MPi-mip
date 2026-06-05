"""Inspect column signatures across multiple bankwise format vintages."""
import sys
import glob
from pathlib import Path
import openpyxl
import xlrd

def get_sigs(filepath):
    ext = Path(filepath).suffix.lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        grid = []
        for r in range(min(8, ws.nrows)):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                v = str(cell.value).replace("\n"," ").strip().lower() if cell.value else ""
                row.append(v)
            grid.append(row)
    else:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        grid = []
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            grid.append([str(v).replace("\n"," ").strip().lower() if v else "" for v in row])

    # Forward fill
    for r in grid:
        last = ""
        for j in range(len(r)):
            if r[j]: last = r[j]
            else: r[j] = last

    max_col = max(len(r) for r in grid)
    sigs = {}
    for j in range(min(max_col, 30)):
        parts = list(dict.fromkeys([r[j] for r in grid if j < len(r) and r[j]]))
        sigs[j+1] = " > ".join(parts[:5])
    return sigs

# Sample files from different eras
files = [
    ("2013", r"D:\MPi-mip\data\raw\rbi_bankwise\2013\ATM290513_APR.xls"),
    ("2017", r"D:\MPi-mip\data\raw\rbi_bankwise\2017\ATM7C0D179A507A4735B4E1343F006DB7CD.XLSX"),
    ("2020", r"D:\MPi-mip\data\raw\rbi_bankwise\2020\ATM02202088363B9FC9FB4B1A83BEB1F67E266ABD.XLS"),
]

for label, fpath in files:
    print(f"\n{'='*80}")
    print(f"FORMAT {label}: {Path(fpath).name}")
    print(f"{'='*80}")
    try:
        sigs = get_sigs(fpath)
        for col, sig in sigs.items():
            if sig:
                print(f"  Col {col:2d}: {sig[:100]}")
    except Exception as e:
        print(f"  ERROR: {e}")
